"""
Helpers for finding, downloading, and previewing attachment files.

The public sites often hide important eligibility details in detail pages or
downloaded files. This module keeps the notebook simple by providing helpers
that:
1. read the latest raw API data,
2. find candidate detail/download links,
3. download the files,
4. extract a short human-readable preview.
"""

from __future__ import annotations

import json
import re
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlparse

import requests

from app.collectors.base import PROJECT_ROOT, ensure_dir

try:
    from docx import Document
except Exception:  # pragma: no cover - optional dependency
    Document = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


ATTACHMENTS_ROOT = PROJECT_ROOT / "data" / "attachments"

DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    )
}

ATTACHMENT_EXTENSIONS = {
    ".pdf",
    ".hwp",
    ".hwpx",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".csv",
    ".txt",
    ".zip",
    ".png",
    ".jpg",
    ".jpeg",
    ".do",
}

ATTACHMENT_KEYWORDS = [
    "attach",
    "attachment",
    "download",
    "file",
    "filedown",
    "getfile",
    "atch",
    "down",
    "pdf",
    "hwp",
    "zip",
    "첨부",
    "다운로드",
    "파일",
    "공고문",
    "신청서",
    "양식",
]

TEXT_HINT_FIELDS = [
    "fileNm",
    "printFileNm",
    "sbmsnDcmntCn",
    "plcyAplyMthdCn",
    "etcMttrCn",
]


class LinkExtractor(HTMLParser):
    """Collect links from an HTML page."""

    def __init__(self) -> None:
        super().__init__()
        self.links: list[dict[str, str]] = []
        self._href: str | None = None
        self._text_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return

        attrs_dict = dict(attrs)
        self._href = attrs_dict.get("href")
        self._text_parts = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() != "a" or self._href is None:
            return

        text = " ".join(part.strip() for part in self._text_parts if part.strip()).strip()
        self.links.append({"href": self._href, "text": text})
        self._href = None
        self._text_parts = []


def _unique_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    unique_values: list[str] = []

    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        unique_values.append(value)

    return unique_values


def sanitize_filename(name: str, default: str = "downloaded_file") -> str:
    cleaned = re.sub(r"[<>:\"/\\|?*]", "_", str(name)).strip()
    cleaned = cleaned.rstrip(". ")
    return cleaned or default


def normalize_probable_url(url: str | None) -> str:
    if not url:
        return ""

    value = str(url).strip()
    if not value:
        return ""

    if value.startswith("//"):
        return f"https:{value}"

    parsed = urlparse(value)
    if parsed.scheme:
        return value

    if value.startswith("www.") or re.match(r"^[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}/", value):
        return f"https://{value}"

    return value


def find_latest_raw_json_file(source_name: str) -> Path | None:
    source_dir = PROJECT_ROOT / "data" / "raw" / source_name
    if not source_dir.exists():
        return None

    json_files = list(source_dir.glob("*.json"))
    if not json_files:
        return None

    return max(json_files, key=lambda path: (path.stat().st_mtime, path.name))


def load_latest_raw_payload(source_name: str) -> tuple[Path | None, Any | None]:
    latest_file = find_latest_raw_json_file(source_name)
    if latest_file is None:
        return None, None

    with latest_file.open("r", encoding="utf-8") as file:
        payload = json.load(file)

    return latest_file, payload


def extract_items(source_name: str, payload: Any) -> list[Any]:
    if payload is None:
        return []

    if source_name == "biz":
        if isinstance(payload, dict):
            if isinstance(payload.get("jsonArray"), list):
                return payload["jsonArray"]
            if isinstance(payload.get("item"), list):
                return payload["item"]
            if isinstance(payload.get("items"), list):
                return payload["items"]
        if isinstance(payload, list):
            return payload
        return []

    if source_name == "kst":
        if isinstance(payload, dict):
            if isinstance(payload.get("data"), list):
                return payload["data"]
            if isinstance(payload.get("items"), list):
                return payload["items"]
        if isinstance(payload, list):
            return payload
        return []

    if source_name == "youth":
        if isinstance(payload, dict):
            result = payload.get("result")
            if isinstance(result, dict):
                if isinstance(result.get("youthPolicyList"), list):
                    return result["youthPolicyList"]
                if isinstance(result.get("items"), list):
                    return result["items"]
                if isinstance(result.get("list"), list):
                    return result["list"]
            if isinstance(result, list):
                return result
            for key in ["data", "items"]:
                if isinstance(payload.get(key), list):
                    return payload[key]
        if isinstance(payload, list):
            return payload
        return []

    if isinstance(payload, list):
        return payload

    return []


def get_latest_items(source_name: str) -> tuple[Path | None, list[Any]]:
    latest_file, payload = load_latest_raw_payload(source_name)
    return latest_file, extract_items(source_name, payload)


def split_inline_file_names(file_text: str | None) -> list[str]:
    if not file_text:
        return []

    parts = [part.strip() for part in str(file_text).split("@")]
    return [part for part in parts if part]


def get_attachment_related_fields(item: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}

    for key, value in item.items():
        lowered_key = key.lower()
        value_text = str(value)

        if any(token in lowered_key for token in ["file", "url", "link", "pdf", "attach"]):
            result[key] = value
            continue

        if key in TEXT_HINT_FIELDS:
            result[key] = value
            continue

        if "http" in value_text.lower():
            result[key] = value

    return result


def collect_candidate_page_urls(source_name: str, item: dict[str, Any]) -> list[str]:
    candidate_fields_by_source = {
        "biz": ["pblancUrl", "rceptEngnHmpgUrl", "flpthNm", "printFlpthNm"],
        "kst": ["detl_pg_url", "biz_gdnc_url", "biz_aply_url"],
        "youth": ["refUrlAddr1", "refUrlAddr2", "aplyUrlAddr"],
    }

    urls: list[str] = []
    for field_name in candidate_fields_by_source.get(source_name, []):
        normalized = normalize_probable_url(item.get(field_name))
        if normalized:
            urls.append(normalized)

    return _unique_keep_order(urls)


def looks_like_downloadable(url: str, text: str = "") -> bool:
    lowered_url = url.lower()
    lowered_text = text.lower()

    parsed_name = Path(urlparse(lowered_url).path).name
    suffix = Path(parsed_name).suffix
    if suffix in ATTACHMENT_EXTENSIONS:
        return True

    return any(keyword.lower() in lowered_url or keyword.lower() in lowered_text for keyword in ATTACHMENT_KEYWORDS)


def fetch_text_response(url: str, timeout: int = 30) -> str:
    response = requests.get(url, headers=DEFAULT_HEADERS, timeout=timeout)
    response.raise_for_status()
    response.encoding = response.encoding or response.apparent_encoding or "utf-8"
    return response.text


def extract_attachment_links_from_html(html_text: str, base_url: str) -> list[dict[str, str]]:
    parser = LinkExtractor()
    parser.feed(html_text)

    candidates: list[dict[str, str]] = []
    for link in parser.links:
        href = normalize_probable_url(link.get("href"))
        if not href:
            continue

        absolute_url = urljoin(base_url, href)
        text = link.get("text", "")
        if not looks_like_downloadable(absolute_url, text):
            continue

        candidates.append(
            {
                "url": absolute_url,
                "text": text,
                "discovered_from": base_url,
            }
        )

    unique_candidates: list[dict[str, str]] = []
    seen_urls: set[str] = set()
    for candidate in candidates:
        candidate_url = candidate["url"]
        if candidate_url in seen_urls:
            continue
        seen_urls.add(candidate_url)
        unique_candidates.append(candidate)

    return unique_candidates


def inspect_attachment_sources(
    source_name: str,
    item: dict[str, Any],
    max_pages: int = 3,
    timeout: int = 30,
) -> dict[str, Any]:
    page_urls = collect_candidate_page_urls(source_name, item)
    inline_file_names = split_inline_file_names(item.get("fileNm")) + split_inline_file_names(item.get("printFileNm"))

    page_results: list[dict[str, Any]] = []
    all_candidates: list[dict[str, str]] = []

    for page_url in page_urls[:max_pages]:
        try:
            html_text = fetch_text_response(page_url, timeout=timeout)
            candidates = extract_attachment_links_from_html(html_text, base_url=page_url)
            page_results.append(
                {
                    "page_url": page_url,
                    "success": True,
                    "candidate_count": len(candidates),
                }
            )
            all_candidates.extend(candidates)
        except Exception as exc:
            page_results.append(
                {
                    "page_url": page_url,
                    "success": False,
                    "candidate_count": 0,
                    "error": str(exc),
                }
            )

    for page_url in page_urls:
        if looks_like_downloadable(page_url):
            all_candidates.append({"url": page_url, "text": "", "discovered_from": "raw_field"})

    unique_candidates: list[dict[str, str]] = []
    seen_urls: set[str] = set()
    for candidate in all_candidates:
        candidate_url = candidate["url"]
        if candidate_url in seen_urls:
            continue
        seen_urls.add(candidate_url)
        unique_candidates.append(candidate)

    return {
        "inline_file_names": inline_file_names,
        "page_urls": page_urls,
        "page_results": page_results,
        "attachment_candidates": unique_candidates,
    }


def _guess_extension_from_content_type(content_type: str) -> str:
    lowered = content_type.lower()
    mapping = {
        "application/pdf": ".pdf",
        "application/zip": ".zip",
        "application/x-zip-compressed": ".zip",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/msword": ".doc",
        "text/plain": ".txt",
        "text/html": ".html",
        "application/json": ".json",
    }

    for key, extension in mapping.items():
        if key in lowered:
            return extension

    return ""


def _filename_from_content_disposition(header_value: str) -> str:
    if not header_value:
        return ""

    match = re.search(r"filename\*=UTF-8''([^;]+)", header_value, flags=re.IGNORECASE)
    if match:
        return sanitize_filename(unquote(match.group(1)))

    match = re.search(r'filename="?([^";]+)"?', header_value, flags=re.IGNORECASE)
    if match:
        return sanitize_filename(unquote(match.group(1)))

    return ""


def _guess_filename_from_response(url: str, response: requests.Response, filename_hint: str = "") -> str:
    content_type = response.headers.get("content-type", "")
    disposition_name = _filename_from_content_disposition(response.headers.get("content-disposition", ""))
    if disposition_name:
        return disposition_name

    parsed = urlparse(url)
    url_name = sanitize_filename(unquote(Path(parsed.path).name))
    if url_name and "." in url_name:
        return url_name

    hint_name = sanitize_filename(filename_hint)
    extension = _guess_extension_from_content_type(content_type)

    if hint_name and "." in hint_name:
        return hint_name
    if hint_name:
        return f"{hint_name}{extension}" if extension else hint_name
    if url_name:
        return f"{url_name}{extension}" if extension else url_name

    return f"downloaded_file{extension}" if extension else "downloaded_file.bin"


def download_attachment(
    url: str,
    source_name: str,
    item_id: str,
    filename_hint: str = "",
    force: bool = False,
    timeout: int = 60,
) -> dict[str, Any]:
    target_dir = ensure_dir(ATTACHMENTS_ROOT / source_name / sanitize_filename(item_id or "unknown_item"))

    response = requests.get(url, headers=DEFAULT_HEADERS, timeout=timeout)
    response.raise_for_status()

    content_type = response.headers.get("content-type", "")
    filename = _guess_filename_from_response(url, response, filename_hint=filename_hint)
    target_path = target_dir / filename

    if target_path.exists() and not force:
        return {
            "url": url,
            "path": target_path,
            "content_type": content_type,
            "downloaded": False,
            "message": "already exists",
        }

    with target_path.open("wb") as file:
        file.write(response.content)

    return {
        "url": url,
        "path": target_path,
        "content_type": content_type,
        "downloaded": True,
        "message": "downloaded",
    }


def strip_html_tags(text: str) -> str:
    without_script = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.IGNORECASE)
    without_style = re.sub(r"<style[\s\S]*?</style>", " ", without_script, flags=re.IGNORECASE)
    without_tags = re.sub(r"<[^>]+>", " ", without_style)
    without_entities = (
        without_tags.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
    )
    normalized = re.sub(r"\s+", " ", without_entities)
    return normalized.strip()


def _repair_common_mojibake(text: str) -> str:
    """
    Some Korean government pages/files arrive as UTF-8 text that was decoded
    once with the wrong single-byte encoding. This repairs the common pattern
    only when it is very likely to help.
    """
    suspicious_count = sum(text.count(marker) for marker in ["ì", "ë", "ê", "í", "ð"])
    if suspicious_count < 3:
        return text

    repaired = ""
    for encoding in ["latin1", "cp1252"]:
        try:
            repaired = text.encode(encoding).decode("utf-8")
            break
        except UnicodeError:
            continue

    if not repaired:
        return text

    return repaired if len(repaired.strip()) >= len(text.strip()) * 0.5 else text


def _preview_text(text: str, max_chars: int = 1200) -> str:
    text = _repair_common_mojibake(text)
    compact = re.sub(r"\s+", " ", text).strip()
    if len(compact) <= max_chars:
        return compact
    return compact[:max_chars] + " ..."


def _read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ["utf-8", "cp949", "euc-kr"]:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _sniff_file_kind(path: Path) -> str:
    head = path.read_bytes()[:4096]
    stripped = head.lstrip().lower()

    if head.startswith(b"%PDF"):
        return "pdf"
    if head.startswith(b"PK\x03\x04"):
        return "zip"
    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "ole"
    if stripped.startswith((b"<!doctype html", b"<html", b"<?xml")) or b"<html" in stripped:
        return "html"

    text_markers = [b"<body", b"<div", b"<table", b"<p", b"<span"]
    if any(marker in stripped for marker in text_markers):
        return "html"

    if b"\x00" not in head:
        return "text"

    return "binary"


def _preview_pdf(path: Path, max_chars: int) -> str:
    if PdfReader is None:
        return "PDF preview needs the pypdf package."

    try:
        reader = PdfReader(str(path))
        page_texts: list[str] = []
        for page in reader.pages[:3]:
            page_texts.append(page.extract_text() or "")
        return _preview_text(" ".join(page_texts), max_chars=max_chars)
    except Exception as exc:
        return f"PDF preview failed: {exc}"


def _preview_zip_family(path: Path, max_chars: int) -> str:
    suffix = path.suffix.lower()

    if suffix == ".docx":
        if Document is None:
            return "DOCX preview needs the python-docx package."
        try:
            document = Document(str(path))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())
            return _preview_text(text, max_chars=max_chars)
        except Exception as exc:
            return f"DOCX preview failed: {exc}"

    if suffix == ".hwpx":
        try:
            collected: list[str] = []
            with zipfile.ZipFile(path) as archive:
                for name in archive.namelist():
                    if name.endswith(".xml"):
                        xml_text = archive.read(name).decode("utf-8", errors="ignore")
                        collected.append(strip_html_tags(xml_text))
            return _preview_text(" ".join(collected), max_chars=max_chars)
        except Exception as exc:
            return f"HWPX preview failed: {exc}"

    if suffix == ".xlsx":
        if load_workbook is None:
            return "XLSX preview needs the openpyxl package."
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            preview_lines = [f"Sheets: {', '.join(sheet_names)}"]
            if sheet_names:
                sheet = workbook[sheet_names[0]]
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    values = [str(value) for value in row if value not in (None, "")]
                    if values:
                        preview_lines.append(" | ".join(values))
                    if row_index >= 5:
                        break
            return _preview_text("\n".join(preview_lines), max_chars=max_chars)
        except Exception as exc:
            return f"XLSX preview failed: {exc}"

    try:
        with zipfile.ZipFile(path) as archive:
            names = [_repair_common_mojibake(name) for name in archive.namelist()]
        if not names:
            return "ZIP file is empty."
        preview = "ZIP file list:\n" + "\n".join(names[:20])
        return _preview_text(preview, max_chars=max_chars)
    except Exception as exc:
        return f"ZIP preview failed: {exc}"


def extract_text_preview(file_path: str | Path, max_chars: int = 1200) -> str:
    path = Path(file_path)
    suffix = path.suffix.lower()
    file_kind = _sniff_file_kind(path)

    if suffix == ".pdf" or file_kind == "pdf":
        return _preview_pdf(path, max_chars=max_chars)

    if suffix in {".docx", ".hwpx", ".xlsx", ".zip"} or file_kind == "zip":
        return _preview_zip_family(path, max_chars=max_chars)

    if suffix == ".hwp" or file_kind == "ole":
        return (
            "HWP/OLE file detected. Automatic text extraction is not supported yet, "
            "but the file was downloaded and should be checked manually."
        )

    if suffix in {".html", ".htm", ".do"} or file_kind == "html":
        html_text = _read_text_with_fallback(path)
        return _preview_text(strip_html_tags(html_text), max_chars=max_chars)

    if suffix in {".txt", ".csv", ".json", ".xml"} or file_kind == "text":
        return _preview_text(_read_text_with_fallback(path), max_chars=max_chars)

    return (
        f"{suffix or 'unknown'} file was downloaded, but automatic preview is not supported yet. "
        "Open the local file if this attachment looks important."
    )


def download_attachment_candidates(
    candidates: list[dict[str, str]],
    source_name: str,
    item_id: str,
    max_files: int = 3,
    force: bool = False,
    timeout: int = 60,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []

    for candidate in candidates[:max_files]:
        try:
            downloaded = download_attachment(
                url=candidate["url"],
                source_name=source_name,
                item_id=item_id,
                filename_hint=candidate.get("text", ""),
                force=force,
                timeout=timeout,
            )
            downloaded["preview_text"] = extract_text_preview(downloaded["path"])
            downloaded["discovered_from"] = candidate.get("discovered_from", "")
            downloaded["link_text"] = candidate.get("text", "")
            results.append(downloaded)
        except Exception as exc:
            results.append(
                {
                    "url": candidate.get("url", ""),
                    "path": None,
                    "downloaded": False,
                    "content_type": "",
                    "message": str(exc),
                    "preview_text": "",
                    "discovered_from": candidate.get("discovered_from", ""),
                    "link_text": candidate.get("text", ""),
                }
            )

    return results
